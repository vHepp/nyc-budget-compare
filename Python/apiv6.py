import argparse
# import os
# import sys
from pathlib import Path
import pandas as pd
import numpy as np
from tqdm.auto import tqdm
from sodapy import Socrata

from datetime import date

APP_KEY = "t2GPRE1GXlktxvqOIFMelPIvT"

version = "6"

client = Socrata(
    "data.cityofnewyork.us",
    APP_KEY,
    timeout=180
)

RELEASE_NAMES = ["prelim", "exec", "adopt"]

ROOT = Path.cwd()
DATA_CACHE = ROOT / "data_cache"


EXPENSE_DATASET_ID = "mwzb-yiwb"

def main(args):
    DATA_CACHE.mkdir(parents=True, exist_ok=True)
    
    print(f"Including contracts: {args.include_contracts}")
    
    
    print("Checking for next fiscal year...")

    ## Check Fiscal Year    
    
    planned_fy = args.fiscal_year
    
    if not planned_fy:
        # Fetch the 2 most recent unique fiscal years present in the dataset
        year_query = "SELECT fiscal_year GROUP BY fiscal_year ORDER BY fiscal_year DESC LIMIT 2"
        year_json = client.get(EXPENSE_DATASET_ID, query=year_query)

        
        # Extract the years into a list, e.g., ['2026', '2025']
        planned_fy = [row['fiscal_year'] for row in year_json][0]

    planned_fy_name = f"FY{int(planned_fy) % 100}"

    current_fy = int(planned_fy) - 1
    current_fy_name = f"FY{int(current_fy) % 100}"

    where_string = f"fiscal_year IN ({planned_fy})"

    orderby_arr = [
        "agency_number", # 2: 
        "unit_appropriation_number", # 4: 
        "responsibility_center_code", # 23: 
        "budget_code_number", # 6: 
        "object_class_number", # 8: 
        "object_code", # 10: 
        "publication_date DESC"
    ]

    orderby_string = ", ".join(orderby_arr)

     
    # Query Data if necessary
    
    data = []
    offset = 0

    limit = 1000

    try:
        print(f"Attempting to load cached data for fiscal year {planned_fy}...")
        df = pd.read_csv(DATA_CACHE / f"backup_{planned_fy}_v{version}.csv")
    except Exception as e:
        # print(e)
        print(f"No cached data found. Querying for fiscal year {planned_fy}...\n")
        while True:
            print(f"Fetching rows starting at offset: {offset}")

            temp = client.get(
                dataset_identifier=EXPENSE_DATASET_ID,
                # select=select_string,
                where=where_string,
                # group=groupby_string,
                order=orderby_string,
                # order=":id",
                limit=limit,
                offset=offset
            )
            
            if not temp:
                break
            
            # print(temp[0])
            
            data.extend(temp)
            
            offset += limit
            
            # break
                
        print(f"End of query. Successfully fetched {len(data)} total rows.")
        
        len(data)
        
        # Convert to pandas DataFrame
        df = pd.DataFrame.from_records(data)
        # df.drop(columns=["financial_plan_savings_flag"], inplace=True)
        
        df.to_csv(DATA_CACHE / f"backup_{planned_fy}_v{version}.csv", index=False)
        df = pd.read_csv(DATA_CACHE / f"backup_{planned_fy}_v{version}.csv")

    
    len(df)


    # rename columns base on target year
    
    str_sum_current_adopted_amount = f"{current_fy_name}_adopted_amount"
    str_sum_current_modified_amount = f"{current_fy_name}_current_modified_amount"
    str_sum_financial_plan_amount = f"{planned_fy_name}_financial_plan_amount"

    str_sum_current_adopted_position = f"{current_fy_name}_adopted_position"
    str_sum_current_modified_position = f"{current_fy_name}_current_modified_position"
    str_sum_financial_plan_position = f"{planned_fy_name}_financial_plan_position"

    str_sum_current_adopted_number_of_contracts = f"{current_fy_name}_adopted_number_of_contracts"
    str_sum_current_modified_number_of_contracts = f"{current_fy_name}_current_modified_number_of_contracts"
    str_sum_financial_plan_number_of_contracts = f"{planned_fy_name}_financial_plan_number_of_contracts"    
     
    rename_cols = {
        "adopted_budget_amount": f"{str_sum_current_adopted_amount}",
        "current_modified_budget_amount": f"{str_sum_current_modified_amount}",
        "financial_plan_amount": f"{str_sum_financial_plan_amount}",
        "adopted_budget_position": f"{str_sum_current_adopted_position}",
        "current_modified_budget_position": f"{str_sum_current_modified_position}",
        "financial_plan_position": f"{str_sum_financial_plan_position}",
        "adopted_budget_number_of_contracts": f"{str_sum_current_adopted_number_of_contracts}",
        "current_modified_budget_number_of_contracts": f"{str_sum_current_modified_number_of_contracts}",
        "financial_plan_number_of_contracts": f"{str_sum_financial_plan_number_of_contracts}",
        }

    df = df.rename(columns=rename_cols)
    
    numeric_cols = [
        str_sum_current_adopted_amount,
        str_sum_current_modified_amount,
        str_sum_financial_plan_amount,

        str_sum_current_adopted_position,
        str_sum_current_modified_position,
        str_sum_financial_plan_position,

        # str_sum_current_adopted_number_of_contracts,
        # str_sum_current_modified_number_of_contracts,
        # str_sum_financial_plan_number_of_contracts,
    ]
    
    if args.include_contracts:
        numeric_cols.append(str_sum_current_adopted_number_of_contracts)
        numeric_cols.append(str_sum_current_modified_number_of_contracts)
        numeric_cols.append(str_sum_financial_plan_number_of_contracts)


    target_categorical_cols = [
        # "publication_date",
        # "fiscal_year",
        "agency_number",
        "agency_name",
        "unit_appropriation_number",
        "unit_appropriation_name",
        "responsibility_center_code",
        "responsibility_center_name",
        "budget_code_number",
        "budget_code_name",
        "object_class_number",
        "object_class_name",
        "object_code",
        "object_code_name",
    ]

    master_categorical_cols = [
        "publication_date",
        "fiscal_year",
        "agency_number",
        "agency_name",
        "unit_appropriation_number",
        "unit_appropriation_name",
        "responsibility_center_code",
        "responsibility_center_name",
        "budget_code_number",
        "budget_code_name",
        "object_class_number",
        "object_class_name",
        "object_code",
        "object_code_name",
    ]

    all_categorical_cols = [
        "publication_date",
        "fiscal_year",
        "agency_number",
        "agency_name",
        "unit_appropriation_number",
        "unit_appropriation_name",
        "responsibility_center_code",
        "responsibility_center_name",
        "budget_code_number",
        "budget_code_name",
        "object_class_number",
        "object_class_name",
        "object_code",
        "object_code_name",
        "financial_plan_savings_flag",
        "intra_city_purchase_code",
        "personal_service_other_than_personal_service_indicator",
    ]

    
    # Sort and reorder
        
    df = df.sort_values(
        [
            "agency_number",
            "unit_appropriation_number",
            "responsibility_center_code",
            "budget_code_number",
            "object_class_number",
            "object_code",
            "publication_date"
            ],
        ascending=[True, True, True, True, True, True, False]
        ).reset_index(drop=True)

    
    # Reorder columns
    df = df[
        [
        "publication_date", # 0: 
        "fiscal_year", # 1: 
        "agency_number", # 2: 
        "agency_name", # 3: 
        "unit_appropriation_number", # 4: 
        "unit_appropriation_name", # 5: 
        "responsibility_center_code", # 23: 
        "responsibility_center_name", # 24: 
        "budget_code_number", # 6: 
        "budget_code_name", # 7: 
        "object_class_number", # 8: 
        "object_class_name", # 9: 
        "object_code", # 10: 
        "object_code_name", # 11: 
        "intra_city_purchase_code", # 25: 
        "personal_service_other_than_personal_service_indicator", # 12: 
        "financial_plan_savings_flag", # 13: 
        *numeric_cols
        # "adopted_budget_amount", # 14: 
        # "current_modified_budget_amount", # 15: 
        # "financial_plan_amount", # 16: 
        # "adopted_budget_position", # 17: 
        # "current_modified_budget_position", # 18: 
        # "financial_plan_position", # 19: 
        # "adopted_budget_number_of_contracts", # 20: 
        # "current_modified_budget_number_of_contracts", # 21: 
        # "financial_plan_number_of_contracts", # 22: 
        ]
    ]
     
    # Clean
    
    df[all_categorical_cols] = df[all_categorical_cols].astype(str)
    
    
    # Fix attempt 1
    # Moved this above df_pivot = df[master_categorical_cols+numeric_cols].groupby(....
    for cat_col in all_categorical_cols:
        if cat_col in df.columns:
            df[cat_col] = df[cat_col].apply(lambda x: str.upper(x) if type(x) == str else x)
    
    
    df_pivot = df[master_categorical_cols+numeric_cols].groupby(master_categorical_cols, dropna=False).sum().reset_index()    

    df_pivot = df_pivot.sort_values(
        [
            "agency_number",
            "unit_appropriation_number",
            "responsibility_center_code",
            "budget_code_number",
            "object_class_number",
            "object_code",
            "publication_date"
            ],
        ascending=[True, True, True, True, True, True, False]
        ).reset_index(drop=True)
    
    df_pivot[master_categorical_cols] = df_pivot[master_categorical_cols].fillna("(blank)")
    df_pivot[numeric_cols] = df_pivot[numeric_cols].fillna(0)
    


    # Track releases
    pub_dates = df_pivot["publication_date"].sort_values(ascending=True).unique().tolist()[:]

    num_pubs_this_year = len(pub_dates)

    print(f"{num_pubs_this_year} pub_dates in FY {planned_fy}: {pub_dates}")

     
    # Set up base DataFrame and map function
    # 

    
    base_df = df_pivot[target_categorical_cols].drop_duplicates().reset_index(drop=True)
    len(base_df)

     
    # Loop through each release
    # 

    
    # print(pub_dates)

    # print(planned_fy)

    for i, pub_date in enumerate(pub_dates):
        release_name = RELEASE_NAMES[i]
        
        ith_release_df = df[df['publication_date'] == pub_date]
        
        # print(i, release_name)

     
    # Collapse Releases
    # 

    
    collapsed_df = base_df.copy()

    new_numeric_cols = []

    start_of_numerical_cols = len(target_categorical_cols)

    # collapsed_df.head()

    for i, pub_date in enumerate(pub_dates):
        print(f"[{i}/{num_pubs_this_year}] pub_date -- {pub_date} ({RELEASE_NAMES[i]}):")
        
        ith_release_df = df_pivot[df_pivot["publication_date"] == pub_date][target_categorical_cols + numeric_cols].copy()
        
        # print(len(ith_release_df))
        
        # print(i)
        
        # print(ith_release_df.columns[len(categorical_cols):])
        
        if i <= 2:
            
            plan_cols = [str_sum_financial_plan_amount, str_sum_financial_plan_position]
            if args.include_contracts:
                plan_cols.append(str_sum_financial_plan_number_of_contracts)
                
            # for col in [str_sum_financial_plan_amount, str_sum_financial_plan_position, str_sum_financial_plan_number_of_contracts]:
            for col in plan_cols:
                # print(col)
                if col in ith_release_df.columns:
                    ith_release_df = ith_release_df.rename(
                        columns={
                            f"{col}":f"{col}_{RELEASE_NAMES[i]}",
                            }
                    )
                # print()
                # print(ith_release_df.columns)
        
            if i < num_pubs_this_year - 1:
                ith_release_df = ith_release_df.drop(columns=[
                    str_sum_current_adopted_amount,
                    str_sum_current_modified_amount,
                    
                    str_sum_current_adopted_position,
                    str_sum_current_modified_position,
                    
                    # str_sum_current_adopted_number_of_contracts,
                    # str_sum_current_modified_number_of_contracts,
                    ])
                
                if args.include_contracts:
                    ith_release_df = ith_release_df.drop(columns=[
                        str_sum_current_adopted_number_of_contracts,
                        str_sum_current_modified_number_of_contracts,
                        ])
                
            # print(f"Adding new numeric columns:{ith_release_df.columns[start_of_numerical_cols:]}\n")
            new_numeric_cols.extend(ith_release_df.columns[start_of_numerical_cols:])
        else:
            raise Exception(f"Bad indexing, i:{i} >= num_pubs_this_year or i:{i} < 0")
        
        # print(ith_release_df.columns)
        # print()
        
        collapsed_df = collapsed_df.merge(right=ith_release_df,on=target_categorical_cols, how='left')
        
        # break

    # collapsed_df = collapsed_df.reset_index(drop=True)


    # print(len(base_df))
    # print(len(collapsed_df))


    collapsed_df.columns

    
    new_numeric_cols

    
    collapsed_df

    
    collapsed_df[collapsed_df.duplicated(keep=False)]

     
    # Object Code Level
    # 

    
    Object_Code_cols = [
        "agency_number",
        "agency_name",
        "unit_appropriation_number",
        "unit_appropriation_name",
        "responsibility_center_code",
        "responsibility_center_name",
        "budget_code_number",
        "budget_code_name",
        "object_class_number",
        "object_class_name",
        "object_code",
        "object_code_name",
        # "financial_plan_savings_flag",
    ]

    main_cols = [
        # 'adopted_budget_amount',
        # 'current_modified_budget_amount',
        f'{str_sum_current_adopted_amount}',
        f'{str_sum_current_modified_amount}',
        f'{str_sum_financial_plan_amount}_{RELEASE_NAMES[i]}',
    ]

    object_code_collapsed_df = collapsed_df.groupby(Object_Code_cols,dropna=False).sum().reset_index()

    object_code_collapsed_df = object_code_collapsed_df[Object_Code_cols + new_numeric_cols]

    object_code_collapsed_df[Object_Code_cols + main_cols]


    
    # object_code_collapsed_df[object_code_collapsed_df.duplicated(keep=False)]

     
    # Object Class Level
    # 

    
    Object_Class_cols = [
        "agency_number",
        "agency_name",
        "unit_appropriation_number",
        "unit_appropriation_name",
        "responsibility_center_code",
        "responsibility_center_name",
        "budget_code_number",
        "budget_code_name",
        "object_class_number",
        "object_class_name",
        # "object_code",
        # "object_code_name",
        # "financial_plan_savings_flag",
    ]

    object_class_collapsed_df = collapsed_df.groupby(Object_Class_cols,dropna=False).sum().reset_index()

    object_class_collapsed_df = object_class_collapsed_df[Object_Class_cols + new_numeric_cols]

    object_class_collapsed_df[Object_Class_cols + main_cols]


    
    # object_code_collapsed_df[object_code_collapsed_df.duplicated(keep=False)]

     
    # Budget Code Level
    # 

    
    Budget_Code_cols = [
        "agency_number",
        "agency_name",
        "unit_appropriation_number",
        "unit_appropriation_name",
        "responsibility_center_code",
        "responsibility_center_name",
        "budget_code_number",
        "budget_code_name",
        # "object_class_number",
        # "object_class_name",
        # "object_code",
        # "object_code_name",
        # "financial_plan_savings_flag",
    ]

    budget_code_collapsed_df = collapsed_df[Budget_Code_cols + new_numeric_cols].groupby(Budget_Code_cols,dropna=False).sum().reset_index()

    # budget_code_collapsed_df = budget_code_collapsed_df[Budget_Code_cols + new_numeric_cols]

    # budget_code_collapsed_df[Budget_Code_cols + main_cols]

    budget_code_collapsed_df

    
    # budget_code_collapsed_df[budget_code_collapsed_df.duplicated(keep=False)]

     
    # Responsibility Center Level
    # 

    
    RC_cols = [
        "agency_number",
        "agency_name",
        "unit_appropriation_number",
        "unit_appropriation_name",
        "responsibility_center_code",
        "responsibility_center_name",
        # "budget_code_number",
        # "budget_code_name",
        # "object_class_number",
        # "object_class_name",
        # "object_code",
        # "object_code_name",
        # "financial_plan_savings_flag",
    ]

    RC_collapsed_df = collapsed_df[RC_cols + new_numeric_cols].groupby(RC_cols,dropna=False).sum().reset_index()

    # RC_collapsed_df = RC_collapsed_df[RC_cols + new_numeric_cols]

    RC_collapsed_df
    # RC_collapsed_df[RC_cols + main_cols]
    # RC_collapsed_df[RC_collapsed_df['agency_name'] == "DEPARTMENT OF EDUCATION"][RC_cols + main_cols]

    
    # RC_collapsed_df[RC_collapsed_df.duplicated(keep=False)]

     
    # Unit of Appropriation Level
    # 

    
    UoA_cols = [
        "agency_number",
        "agency_name",
        "unit_appropriation_number",
        "unit_appropriation_name",
        # "responsibility_center_code",
        # "responsibility_center_name",
        # "budget_code_number",
        # "budget_code_name",
        # "object_class_number",
        # "object_class_name",
        # "object_code",
        # "object_code_name",
        # "financial_plan_savings_flag",
    ]

    UoA_collapsed_df = collapsed_df[UoA_cols + new_numeric_cols].groupby(UoA_cols,dropna=False).sum().reset_index()

    # UoA_collapsed_df = UoA_collapsed_df[UoA_cols + new_numeric_cols]

    UoA_collapsed_df
    # UoA_collapsed_df[UoA_cols + main_cols]
    # UoA_collapsed_df[UoA_collapsed_df['agency_name'] == "DEPARTMENT OF EDUCATION"][UoA_cols + main_cols]

    
    # UoA_collapsed_df[UoA_collapsed_df.duplicated(keep=False)]

     
    # Agency Level
    # 

    
    Agency_cols = [
        "agency_number",
        "agency_name",
        # "unit_appropriation_number",
        # "unit_appropriation_name",
        # "responsibility_center_code",
        # "responsibility_center_name",
        # "budget_code_number",
        # "budget_code_name",
        # "object_class_number",
        # "object_class_name",
        # "object_code",
        # "object_code_name",
        # "financial_plan_savings_flag",
    ]

    Agency_collapsed_df = collapsed_df[Agency_cols + new_numeric_cols].groupby(Agency_cols,dropna=False).sum().reset_index()

    # Agency_collapsed_df = Agency_collapsed_df[Agency_cols + new_numeric_cols]

    # Agency_collapsed_df[Agency_cols + main_cols]
    Agency_collapsed_df

     
    # Write to excel
    # 

    
    
    def add_diff_cols(df):
        diff_df = df.copy()
        
        # --- NEW HELPER FUNCTION TO PREVENT CRASHES ---
        def safe_pct_change(change_series, base_series):
            # change / base, replacing division-by-zero infinities with NaN
            return change_series.div(base_series).replace([np.inf, -np.inf], np.nan)
        # ----------------------------------------------

        for i in range(0,num_pubs_this_year):
            i_amount_name = f"{str_sum_financial_plan_amount}_{RELEASE_NAMES[i]}"
            i_position_name = f"{str_sum_financial_plan_position}_{RELEASE_NAMES[i]}"
            i_number_of_contracts_name = f"{str_sum_financial_plan_number_of_contracts}_{RELEASE_NAMES[i]}"

            # Amount Changes
            amount_change_prev_adopt_name = f"{str_sum_financial_plan_amount}_change_{RELEASE_NAMES[i]}-{current_fy_name}_adopt"
            amount_change_percent_prev_adopt_name = f"{str_sum_financial_plan_amount}_%change_{RELEASE_NAMES[i]}-{current_fy_name}_adopt"
            amount_change_prev_modified_name = f"{str_sum_financial_plan_amount}_change_{RELEASE_NAMES[i]}-{current_fy_name}_modified"
            amount_change_percent_prev_modified_name = f"{str_sum_financial_plan_amount}_%change_{RELEASE_NAMES[i]}-{current_fy_name}_modified"

            diff_df[amount_change_prev_adopt_name]= df[i_amount_name] - df[str_sum_current_adopted_amount]
            # FIXED: Safe divide by the OLD amount (Adopted)
            diff_df[amount_change_percent_prev_adopt_name]= safe_pct_change(diff_df[amount_change_prev_adopt_name], df[str_sum_current_adopted_amount])

            diff_df[amount_change_prev_modified_name] = df[i_amount_name] - df[str_sum_current_modified_amount]
            # FIXED: Safe divide by the OLD amount (Modified)
            diff_df[amount_change_percent_prev_modified_name] = safe_pct_change(diff_df[amount_change_prev_modified_name], df[str_sum_current_modified_amount])


            # Position Changes
            position_change_prev_adopt_name = f"{str_sum_financial_plan_position}_change_{RELEASE_NAMES[i]}-{current_fy_name}_adopt"
            # position_change_percent_prev_adopt_name = f"{str_sum_financial_plan_position}_%change_{RELEASE_NAMES[i]}-{current_fy_name}_adopt"
            position_change_prev_modified_name = f"{str_sum_financial_plan_position}_change_{RELEASE_NAMES[i]}-{current_fy_name}_modified"
            # position_change_percent_prev_modified_name = f"{str_sum_financial_plan_position}_%change_{RELEASE_NAMES[i]}-{current_fy_name}_modified"

            diff_df[position_change_prev_adopt_name] = df[i_position_name] - df[str_sum_current_adopted_position]
            # FIXED
            # diff_df[position_change_percent_prev_adopt_name] = safe_pct_change(diff_df[position_change_prev_adopt_name], df[str_sum_current_adopted_position])
            
            diff_df[position_change_prev_modified_name] = df[i_position_name] - df[str_sum_current_modified_position]
            # FIXED
            # diff_df[position_change_percent_prev_modified_name] = safe_pct_change(diff_df[position_change_prev_modified_name], df[str_sum_current_modified_position])
            
            
            # Contract Changes
            number_of_contracts_change_prev_adopt_name = f"{str_sum_financial_plan_number_of_contracts}_change_{RELEASE_NAMES[i]}-{current_fy_name}_adopt"
            # number_of_contracts_change_percent_prev_adopt_name = f"{str_sum_financial_plan_number_of_contracts}_%change_{RELEASE_NAMES[i]}-{current_fy_name}_adopt"
            number_of_contracts_change_prev_modified_name = f"{str_sum_financial_plan_number_of_contracts}_change_{RELEASE_NAMES[i]}-{current_fy_name}_modified"
            # number_of_contracts_change_percent_prev_modified_name = f"{str_sum_financial_plan_number_of_contracts}_%change_{RELEASE_NAMES[i]}-{current_fy_name}_modified"

            if args.include_contracts:

                diff_df[number_of_contracts_change_prev_adopt_name] = df[i_number_of_contracts_name] - df[str_sum_current_adopted_number_of_contracts]
                # FIXED
                # diff_df[number_of_contracts_change_percent_prev_adopt_name] = safe_pct_change(diff_df[number_of_contracts_change_prev_adopt_name], df[str_sum_current_adopted_number_of_contracts])
                
                diff_df[number_of_contracts_change_prev_modified_name] = df[i_number_of_contracts_name] - df[str_sum_current_modified_number_of_contracts]
                # FIXED
                # diff_df[number_of_contracts_change_percent_prev_modified_name] = safe_pct_change(diff_df[number_of_contracts_change_prev_modified_name], df[str_sum_current_modified_number_of_contracts])
            
            # Release-to-Release comparisons
            for j in range(0,i):
                j_amount_name = f"{str_sum_financial_plan_amount}_{RELEASE_NAMES[j]}"
                j_position_name = f"{str_sum_financial_plan_position}_{RELEASE_NAMES[j]}"
                j_number_of_contracts_name = f"{str_sum_financial_plan_number_of_contracts}_{RELEASE_NAMES[j]}"
                    
                amount_change_name = f"{str_sum_financial_plan_amount}_change_{RELEASE_NAMES[i]}-{RELEASE_NAMES[j]}"
                amount_change_percent_name = f"{str_sum_financial_plan_amount}_%change_{RELEASE_NAMES[i]}-{RELEASE_NAMES[j]}"
                    
                diff_df[amount_change_name] = df[i_amount_name] - df[j_amount_name]
                # FIXED: Divide by 'j' (the older release), not 'i'
                diff_df[amount_change_percent_name] = safe_pct_change(diff_df[amount_change_name], df[j_amount_name])
                
                position_change_name = f"{str_sum_financial_plan_position}_change_{RELEASE_NAMES[i]}-{RELEASE_NAMES[j]}"
                # position_change_percent_name = f"{str_sum_financial_plan_position}_%change_{RELEASE_NAMES[i]}-{RELEASE_NAMES[j]}"
                
                diff_df[position_change_name] = df[i_position_name] - df[j_position_name]
                # FIXED
                # diff_df[position_change_percent_name] = safe_pct_change(diff_df[position_change_name], df[j_position_name])
                
                number_of_contracts_change_name = f"{str_sum_financial_plan_number_of_contracts}_change_{RELEASE_NAMES[i]}-{RELEASE_NAMES[j]}"
                # number_of_contracts_change_percent_name = f"{str_sum_financial_plan_number_of_contracts}_%change_{RELEASE_NAMES[i]}-{RELEASE_NAMES[j]}"
                
                
                if args.include_contracts:
                    diff_df[number_of_contracts_change_name] = df[i_number_of_contracts_name] - df[j_number_of_contracts_name]
                    # FIXED
                    # diff_df[number_of_contracts_change_percent_name] = safe_pct_change(diff_df[number_of_contracts_change_name], df[j_number_of_contracts_name])
            
        return diff_df
    # add_diff_cols(df)

    
    from openpyxl.utils import get_column_letter

    output_filename = f"final_export_{planned_fy}.xlsx"

    # 1. Group your sheets into a dictionary to easily loop through them
    sheets_to_write = {
        "Raw": df_pivot,
        "Collapsed": collapsed_df,
        "Object Code": object_code_collapsed_df,
        "Object Class": object_class_collapsed_df,
        "Budget Code": budget_code_collapsed_df,
        "Responsibility Center": RC_collapsed_df,
        "Unit of Appropriation": UoA_collapsed_df,
        "Agency": Agency_collapsed_df
    }

    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        for k, (sheet_name, df) in enumerate(tqdm(sheets_to_write.items())):
            
            # print(k)
            
            if k > 0:
                df = add_diff_cols(df)
            
            # A. Write the dataframe to the current sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            
            # Formating
            if sheet_name != "Raw":
                
                # B. Access the active openpyxl worksheet object pandas just generated
                worksheet = writer.sheets[sheet_name]
                
                # C. Loop through the columns of the DataFrame
                # (enumerate start=1 because Excel columns are 1-indexed: A=1, B=2...)
                for col_idx, col_name in enumerate(df.columns, start=1):
                    
                    # Check if this column is one of your targets
                    if col_name not in all_categorical_cols:
                        
                        # Assign formatting style: Currency vs Headcount Positions
                        # if "position" in col_name.lower() or "contract" in col_name.lower():
                        #     num_format = '#,##0'   
                        # elif "%" in col_name.lower():
                        #     num_format = '#,##0'   
                        # elif "amount" in col_name.lower():
                        #     num_format = '$#,##0'  
                        # else:
                        #     num_format = '#,##0'   
                        
                        # --- FORMATTING BLOCK ---
                        if "position" in col_name.lower() or "contract" in col_name.lower():
                            num_format = '#,##0'   
                        elif "%" in col_name.lower() or "percent" in col_name.lower():
                            num_format = '0.0%'  
                        elif "amount" in col_name.lower():
                            num_format = '$#,##0'  
                        else:
                            num_format = '#,##0'   
                        # -----------------------------------
                        
                        # Apply the format to every cell in this column, skipping the header (row 1)
                        for row in range(2, worksheet.max_row + 1):
                            worksheet.cell(row=row, column=col_idx).number_format = num_format
                        
                        # BONUS: Set a wider column width so Excel doesn't clip numbers into '###'
                        col_letter = get_column_letter(col_idx)
                        worksheet.column_dimensions[col_letter].width = 22

    print(f"Successfully saved and formatted multiple sheets to {output_filename}!")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="NYC Expense Budget Pipeline")

    parser.add_argument(
        "--fiscal_year",
        type=str,
        default= 2027,
        help="Target fiscal year for analysis. (Default: 2027)",
    )
    
    parser.add_argument(
        "--include_contracts",
        action="store_true",
        help="Includes number of contracts columns in analysis.",
    )

    # parser.add_argument(
    #     "--mode",
    #     type=str,
    #     choices=["train", "inference", "both"],
    #     required=True,
    #     help="Choose whether to train a new model, run inference, or do both. ['train', 'inference', 'both'].",
    # )
    # parser.add_argument(
    #     "--model_name",
    #     type=str,
    #     # required=True,
    #     help="Base name for saving/loading the models.",
    # )

    # parser.add_argument(
    #     "--submission_name",
    #     type=str,
    #     default=None,
    #     help="Name of csv file containing inference results. Defaults to submission_cellpose.",
    # )

    # parser.add_argument(
    #     "--epochs",
    #     type=int,
    #     default=50,
    #     help="Number of epochs to train each fold for (default 50).",
    # )

    # parser.add_argument(
    #     "--batch_size",
    #     type=int,
    #     default=1,
    #     help="Size of training batches (default 1).",
    # )

    # parser.add_argument(
    #     "--val_size",
    #     type=float,
    #     default=0.2,
    #     help="Proportion of sessions to use for validation in 'single' mode (default 0.2).",
    # )

    # parser.add_argument(
    #     "--subset",
    #     type=float,
    #     default=1.0,
    #     help="Portion of training sessions to use. Lower values mean training on smaller subsets and faster prototyping (default 1.0).",
    # )

    # --- NEW HYPERPARAMETERS ---
    # parser.add_argument(
    #     "--stitch_threshold",
    #     type=float,
    #     default=0.5,
    # )

    # parser.add_argument(
    #     "--flow_threshold",
    #     type=float,
    #     default=0.4,
    # )

    # parser.add_argument(
    #     "--cellprob_threshold",
    #     type=float,
    #     default=0.0,
    #     help="Threshold for cell probability. Lower values (e.g., -1.5) expand the predicted masks.",
    # )

    # parser.add_argument(
    #     "--min_size",
    #     type=int,
    #     default=15,
    #     help="Minimum size of cells (in voxels/pixels). Masks smaller than this are deleted.",
    # )

    # ---------------------------

    args = parser.parse_args()

    main(args)